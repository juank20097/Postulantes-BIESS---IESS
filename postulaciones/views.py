import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.db.models import Count
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from xhtml2pdf import pisa
from .models import (
    Postulante, FamiliarIESS, FormacionAcademica,
    ExperienciaProfesional, Capacitacion, Publicacion, Inhabilidades
)
from usuarios.email_service import notificar_cambio_estado

MENSAJE_INHABILITADO = (
    'Su postulación NO puede continuar, por cuanto se encuentra dentro de las '
    'prohibiciones e inhabilidades (art. 13, Resolución No. C.D. 701).'
)

BLOQUEOS_INHABILIDADES = {
    'p1_goce_derechos':         False,
    'p2_inhabilitado_comercio': True,
    'p3_mora_obligaciones':     True,
    'p4_vinculo_financiero':    True,
    'p5_funcionario_iess':      True,
    'p6_interes_aseguradoras':  True,
    'p7_removido_organismo':    True,
    'p8_sri':                   True,
    'p9_castigo_financiero':    True,
    'p10_litigio_iess':         True,
    'p11_procesado_corrupcion': True,
    'p12_contraloria':          True,
    'p13_uafe':                 True,
}

MAX_2MB = 2 * 1024 * 1024  # 2 MB en bytes


def get_or_create_postulante(user):
    postulante, _ = Postulante.objects.get_or_create(usuario=user)
    return postulante


@login_required
def paso_info_personal(request):
    postulante = get_or_create_postulante(request.user)
    if request.method == 'POST':
        postulante.nombres          = request.POST.get('nombres', '')
        postulante.apellidos        = request.POST.get('apellidos', '')
        postulante.genero           = request.POST.get('genero', '')
        postulante.fecha_nacimiento = request.POST.get('fecha_nacimiento') or None
        postulante.nacionalidad     = request.POST.get('nacionalidad', '')
        postulante.estado_civil     = request.POST.get('estado_civil', '')
        postulante.conyuge_nombres  = request.POST.get('conyuge_nombres', '')
        postulante.conyuge_cedula   = request.POST.get('conyuge_cedula', '')
        postulante.save()
        messages.success(request, 'Información personal guardada.')
        return redirect('paso_domicilio')
    return render(request, 'postulaciones/paso_info_personal.html', {
        'postulante': postulante, 'paso': 1, 'total_pasos': 8
    })


@login_required
def paso_domicilio(request):
    postulante = get_or_create_postulante(request.user)
    if request.method == 'POST':
        postulante.pais               = request.POST.get('pais', '')
        postulante.provincia          = request.POST.get('provincia', '')
        postulante.ciudad             = request.POST.get('ciudad', '')
        postulante.calle_principal    = request.POST.get('calle_principal', '')
        postulante.numero             = request.POST.get('numero', '')
        postulante.calle_secundaria   = request.POST.get('calle_secundaria', '')
        postulante.sector_domicilio   = request.POST.get('sector_domicilio', '')
        postulante.referencia         = request.POST.get('referencia', '')
        postulante.telefono_celular   = request.POST.get('telefono_celular', '')
        postulante.telefono_domicilio = request.POST.get('telefono_domicilio', '')
        postulante.email_secundario   = request.POST.get('email_secundario', '')
        postulante.save()
        messages.success(request, 'Domicilio guardado.')
        return redirect('paso_sector')
    return render(request, 'postulaciones/paso_domicilio.html', {
        'postulante': postulante, 'paso': 2, 'total_pasos': 8
    })


@login_required
def paso_sector(request):
    postulante = get_or_create_postulante(request.user)
    if request.method == 'POST':
        postulante.sector = request.POST.get('sector', '')
        tiene_org = request.POST.get('tiene_organizacion')
        if tiene_org == 'True':
            postulante.tiene_organizacion = True
        elif tiene_org == 'False':
            postulante.tiene_organizacion = False
        else:
            postulante.tiene_organizacion = None
        postulante.nombre_organizacion = request.POST.get('nombre_organizacion', '')
        if 'doc_organizacion' in request.FILES:
            postulante.doc_organizacion = request.FILES['doc_organizacion']
        postulante.save()
        messages.success(request, 'Sector guardado.')
        return redirect('paso_familiares')
    return render(request, 'postulaciones/paso_sector.html', {
        'postulante': postulante, 'paso': 3, 'total_pasos': 8
    })


@login_required
def paso_familiares(request):
    postulante = get_or_create_postulante(request.user)
    if request.method == 'POST':
        ids = set()
        for key in request.POST:
            if key.startswith('familiar_nombres_'):
                ids.add(key.replace('familiar_nombres_', ''))
        for fid in ids:
            nombres     = request.POST.get(f'familiar_nombres_{fid}', '').strip()
            parentesco  = request.POST.get(f'familiar_parentesco_{fid}', '')
            institucion = request.POST.get(f'familiar_institucion_{fid}', '')
            area        = request.POST.get(f'familiar_area_{fid}', '').strip()
            cargo       = request.POST.get(f'familiar_cargo_{fid}', '').strip()
            if nombres and parentesco and institucion and area and cargo:
                FamiliarIESS.objects.create(
                    postulante=postulante, nombres=nombres, parentesco=parentesco,
                    institucion=institucion, area=area, cargo=cargo,
                )
        messages.success(request, 'Familiares guardados.')
        return redirect('paso_formacion')
    familiares = FamiliarIESS.objects.filter(postulante=postulante)
    return render(request, 'postulaciones/paso_familiares.html', {
        'familiares': familiares, 'postulante': postulante, 'paso': 4, 'total_pasos': 8
    })


@login_required
def paso_formacion(request):
    postulante = get_or_create_postulante(request.user)
    if request.method == 'POST':
        ids = set()
        for key in request.POST:
            if key.startswith('formacion_nivel_'):
                ids.add(key.replace('formacion_nivel_', ''))
        for fid in ids:
            nivel          = request.POST.get(f'formacion_nivel_{fid}', '')
            tipo           = request.POST.get(f'formacion_tipo_{fid}', '')
            institucion    = request.POST.get(f'formacion_institucion_{fid}', '').strip()
            area           = request.POST.get(f'formacion_area_{fid}', '')
            titulo         = request.POST.get(f'formacion_titulo_{fid}', '').strip()
            num_senescyt   = request.POST.get(f'formacion_senescyt_{fid}', '').strip()
            fecha_senescyt = request.POST.get(f'formacion_fecha_senescyt_{fid}', '')
            documento      = request.FILES.get(f'formacion_documento_{fid}')
            if nivel and institucion and titulo and num_senescyt and fecha_senescyt and documento:
                FormacionAcademica.objects.create(
                    postulante=postulante, nivel=nivel, tipo=tipo, institucion=institucion,
                    area_estudios=area, titulo=titulo, num_senescyt=num_senescyt,
                    fecha_senescyt=fecha_senescyt, documento=documento,
                )
        messages.success(request, 'Formación académica guardada.')
        return redirect('paso_experiencia')
    formaciones = FormacionAcademica.objects.filter(postulante=postulante)
    return render(request, 'postulaciones/paso_formacion.html', {
        'formaciones': formaciones, 'postulante': postulante, 'paso': 5, 'total_pasos': 8
    })


@login_required
def eliminar_formacion(request, pk):
    postulante = get_or_create_postulante(request.user)
    formacion  = get_object_or_404(FormacionAcademica, pk=pk, postulante=postulante)
    formacion.delete()
    messages.success(request, 'Formación eliminada.')
    return redirect('paso_formacion')


@login_required
def paso_experiencia(request):
    postulante = get_or_create_postulante(request.user)
    if request.method == 'POST':
        ids = set()
        for key in request.POST:
            if key.startswith('exp_tipo_'):
                ids.add(key.replace('exp_tipo_', ''))
        for eid in ids:
            tipo         = request.POST.get(f'exp_tipo_{eid}', '')
            area         = request.POST.get(f'exp_area_{eid}', '')
            cargo        = request.POST.get(f'exp_cargo_{eid}', '').strip()
            institucion  = request.POST.get(f'exp_institucion_{eid}', '').strip()
            fecha_inicio = request.POST.get(f'exp_fecha_inicio_{eid}', '')
            fecha_fin    = request.POST.get(f'exp_fecha_fin_{eid}', '')
            descripcion  = request.POST.get(f'exp_descripcion_{eid}', '').strip()
            documento    = request.FILES.get(f'exp_documento_{eid}')
            if tipo and cargo and institucion and fecha_inicio and fecha_fin and descripcion and documento:
                ExperienciaProfesional.objects.create(
                    postulante=postulante, tipo=tipo, actividades_area=area, cargo=cargo,
                    institucion=institucion, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin,
                    descripcion=descripcion, documento=documento,
                )
        messages.success(request, 'Experiencia profesional guardada.')
        return redirect('paso_capacitacion')
    experiencias = ExperienciaProfesional.objects.filter(postulante=postulante)
    return render(request, 'postulaciones/paso_experiencia.html', {
        'experiencias': experiencias, 'postulante': postulante, 'paso': 6, 'total_pasos': 8
    })


@login_required
def eliminar_experiencia(request, pk):
    postulante  = get_or_create_postulante(request.user)
    experiencia = get_object_or_404(ExperienciaProfesional, pk=pk, postulante=postulante)
    experiencia.delete()
    messages.success(request, 'Experiencia eliminada.')
    return redirect('paso_experiencia')


@login_required
def paso_capacitacion(request):
    postulante = get_or_create_postulante(request.user)
    if request.method == 'POST':
        cap_ids = set()
        for key in request.POST:
            if key.startswith('cap_tipo_'):
                cap_ids.add(key.replace('cap_tipo_', ''))
        for cid in cap_ids:
            tipo         = request.POST.get(f'cap_tipo_{cid}', '')
            nombre       = request.POST.get(f'cap_nombre_{cid}', '').strip()
            institucion  = request.POST.get(f'cap_institucion_{cid}', '').strip()
            fecha_inicio = request.POST.get(f'cap_fecha_inicio_{cid}', '')
            fecha_fin    = request.POST.get(f'cap_fecha_fin_{cid}', '')
            horas        = request.POST.get(f'cap_horas_{cid}', 0)
            documento    = request.FILES.get(f'cap_documento_{cid}')
            if tipo and nombre and institucion and fecha_inicio and fecha_fin and horas and documento:
                Capacitacion.objects.create(
                    postulante=postulante, tipo_evento=tipo, nombre=nombre, institucion=institucion,
                    fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, horas=horas, documento=documento,
                )
        pub_ids = set()
        for key in request.POST:
            if key.startswith('pub_titulo_'):
                pub_ids.add(key.replace('pub_titulo_', ''))
        for pid in pub_ids:
            titulo      = request.POST.get(f'pub_titulo_{pid}', '').strip()
            tipo        = request.POST.get(f'pub_tipo_{pid}', '')
            medio       = request.POST.get(f'pub_medio_{pid}', '').strip()
            fecha       = request.POST.get(f'pub_fecha_{pid}', '')
            relacionado = request.POST.get(f'pub_relacionado_{pid}', 'False') == 'True'
            documento   = request.FILES.get(f'pub_documento_{pid}')
            if titulo and tipo and medio and fecha and documento:
                Publicacion.objects.create(
                    postulante=postulante, titulo=titulo, tipo=tipo, medio=medio,
                    fecha=fecha, relacionado=relacionado, documento=documento,
                )
        messages.success(request, 'Capacitación y publicaciones guardadas.')
        return redirect('paso_inhabilidades')
    capacitaciones = Capacitacion.objects.filter(postulante=postulante)
    publicaciones  = Publicacion.objects.filter(postulante=postulante)
    return render(request, 'postulaciones/paso_capacitacion.html', {
        'capacitaciones': capacitaciones, 'publicaciones': publicaciones,
        'postulante': postulante, 'paso': 7, 'total_pasos': 8
    })


@login_required
def eliminar_capacitacion(request, pk):
    postulante   = get_or_create_postulante(request.user)
    capacitacion = get_object_or_404(Capacitacion, pk=pk, postulante=postulante)
    capacitacion.delete()
    messages.success(request, 'Capacitación eliminada.')
    return redirect('paso_capacitacion')


@login_required
def eliminar_publicacion(request, pk):
    postulante  = get_or_create_postulante(request.user)
    publicacion = get_object_or_404(Publicacion, pk=pk, postulante=postulante)
    publicacion.delete()
    messages.success(request, 'Publicación eliminada.')
    return redirect('paso_capacitacion')


@login_required
def paso_inhabilidades(request):
    postulante     = get_or_create_postulante(request.user)
    inhabilidad, _ = Inhabilidades.objects.get_or_create(postulante=postulante)

    preguntas = [
        ('p1_goce_derechos',        '¿Es ecuatoriano/a y está en pleno goce de los derechos de participación política?'),
        ('p2_inhabilitado_comercio', '¿Se encuentra inhabilitado para ejercer el comercio?'),
        ('p3_mora_obligaciones',     '¿Está en mora, directa o indirectamente, en el pago de sus obligaciones con el Estado o SuperBancos?'),
        ('p4_vinculo_financiero',    '¿Mantiene algún vínculo con instituciones del sistema financiero o de seguros privados?'),
        ('p5_funcionario_iess',      '¿Es funcionario o empleado del IESS o BIESS?'),
        ('p6_interes_aseguradoras',  '¿Mantiene interés propio o representa a terceros en compañías aseguradoras del sistema de seguridad social?'),
        ('p7_removido_organismo',    '¿En los últimos 5 años, ha sido removido por algún Organismo de Control?'),
        ('p8_sri',                   '¿Mantiene obligaciones pendientes con el SRI?'),
        ('p9_castigo_financiero',    '¿Ha incurrido en los últimos 5 años en castigo de obligaciones por institución financiera?'),
        ('p10_litigio_iess',         '¿Actualmente se encuentra litigando en contra del IESS o BIESS?'),
        ('p11_procesado_corrupcion', '¿Se encuentra procesado, enjuiciado o condenado por algún delito de corrupción o crimen organizado?'),
        ('p12_contraloria',          '¿Mantiene en firme algún tipo de responsabilidades por la Contraloría General del Estado?'),
        ('p13_uafe',                 '¿Se encuentra registrado en la base de datos de la UAFE?'),
    ]

    if request.method == 'POST':
        for campo, _ in preguntas:
            valor = request.POST.get(campo)
            if valor is not None:
                setattr(inhabilidad, campo, valor == 'True')
        inhabilidad.p7_institucion = request.POST.get('p7_institucion', '')
        inhabilidad.p7_cargo       = request.POST.get('p7_cargo', '')
        inhabilidad.p7_motivo      = request.POST.get('p7_motivo', '')
        fecha_p7 = request.POST.get('p7_fecha_fin')
        inhabilidad.p7_fecha_fin   = fecha_p7 if fecha_p7 else None
        inhabilidad.save()

        campo_bloqueado = None
        for campo, valor_bloqueo in BLOQUEOS_INHABILIDADES.items():
            respuesta = getattr(inhabilidad, campo, None)
            if respuesta is not None and respuesta == valor_bloqueo:
                campo_bloqueado = campo
                break

        if campo_bloqueado:
            postulante.estado = 'INHABILITADO'
            postulante.save()
            notificar_cambio_estado(postulante)

            valores_actuales = {}
            for campo, _ in preguntas:
                val = getattr(inhabilidad, campo, None)
                valores_actuales[campo] = 'True' if val is True else ('False' if val is False else '')

            return render(request, 'postulaciones/paso_inhabilidades.html', {
                'inhabilidad':           inhabilidad,
                'postulante':            postulante,
                'preguntas':             preguntas,
                'valores_actuales_json': json.dumps(valores_actuales),
                'paso': 8, 'total_pasos': 8,
                'bloqueado':       True,
                'campo_bloqueado': campo_bloqueado,
                'mensaje_bloqueo': MENSAJE_INHABILITADO,
            })

        messages.success(request, 'Inhabilidades guardadas.')
        return redirect('resumen_postulacion')

    valores_actuales = {}
    for campo, _ in preguntas:
        val = getattr(inhabilidad, campo, None)
        valores_actuales[campo] = 'True' if val is True else ('False' if val is False else '')

    return render(request, 'postulaciones/paso_inhabilidades.html', {
        'inhabilidad':           inhabilidad,
        'postulante':            postulante,
        'preguntas':             preguntas,
        'valores_actuales_json': json.dumps(valores_actuales),
        'paso': 8, 'total_pasos': 8,
        'bloqueado': False, 'mensaje_bloqueo': '',
    })


@login_required
def resumen_postulacion(request):
    postulante = get_object_or_404(Postulante, usuario=request.user)
    errores_docs = []

    if request.method == 'POST':
        declaracion = request.FILES.get('declaracion_juramentada')
        constancia  = request.FILES.get('constancia')

        # ── Validar que ambos documentos fueron subidos ───────────────────
        if not declaracion:
            errores_docs.append('Debe subir el documento de Declaración Juramentada.')
        if not constancia:
            errores_docs.append('Debe subir el documento de Constancia.')

        # ── Validar PDF y tamaño máximo 2 MB por documento ───────────────
        if declaracion:
            if not declaracion.name.lower().endswith('.pdf'):
                errores_docs.append('La Declaración Juramentada debe ser un archivo PDF.')
            elif declaracion.size > MAX_2MB:
                errores_docs.append('La Declaración Juramentada no puede superar los 2 MB.')

        if constancia:
            if not constancia.name.lower().endswith('.pdf'):
                errores_docs.append('La Constancia debe ser un archivo PDF.')
            elif constancia.size > MAX_2MB:
                errores_docs.append('La Constancia no puede superar los 2 MB.')

        if errores_docs:
            return render(request, 'postulaciones/resumen.html', {
                'postulante':     postulante,
                'formaciones':    postulante.formaciones.all(),
                'experiencias':   postulante.experiencias.all(),
                'capacitaciones': postulante.capacitaciones.all(),
                'publicaciones':  postulante.publicaciones.all(),
                'familiares':     postulante.familiares.all(),
                'inhabilidades':  getattr(postulante, 'inhabilidades', None),
                'errores_docs':   errores_docs,
                'mostrar_modal':  True,
            })

        # ── Guardar documentos en MinIO a través del modelo ───────────────
        postulante.declaracion_juramentada = declaracion
        postulante.constancia              = constancia
        postulante.estado                  = 'ENVIADO'
        postulante.save()

        notificar_cambio_estado(postulante)
        messages.success(request, f'Postulación {postulante.codigo_unico} enviada correctamente.')
        return redirect('confirmacion_postulacion')

    return render(request, 'postulaciones/resumen.html', {
        'postulante':     postulante,
        'formaciones':    postulante.formaciones.all(),
        'experiencias':   postulante.experiencias.all(),
        'capacitaciones': postulante.capacitaciones.all(),
        'publicaciones':  postulante.publicaciones.all(),
        'familiares':     postulante.familiares.all(),
        'inhabilidades':  getattr(postulante, 'inhabilidades', None),
        'errores_docs':   [],
        'mostrar_modal':  False,
    })


@login_required
def eliminar_familiar(request, pk):
    postulante = get_or_create_postulante(request.user)
    familiar   = get_object_or_404(FamiliarIESS, pk=pk, postulante=postulante)
    familiar.delete()
    messages.success(request, 'Familiar eliminado.')
    return redirect('paso_familiares')


@login_required
def confirmacion_postulacion(request):
    postulante = get_object_or_404(Postulante, usuario=request.user)
    return render(request, 'postulaciones/confirmacion.html', {'postulante': postulante})


@login_required
def descargar_pdf(request):
    postulante     = get_object_or_404(Postulante, usuario=request.user)
    familiares     = postulante.familiares.all()
    formaciones    = postulante.formaciones.all()
    experiencias   = postulante.experiencias.all()
    capacitaciones = postulante.capacitaciones.all()
    publicaciones  = postulante.publicaciones.all()
    inhabilidades  = getattr(postulante, 'inhabilidades', None)

    def si_no(val):
        if val is True: return 'Si'
        elif val is False: return 'No'
        return '-'

    inh_rows = ''
    if inhabilidades:
        for campo, etiqueta in [
            ('p1_goce_derechos','9.1. Ecuatoriano/a en pleno goce de derechos?'),
            ('p2_inhabilitado_comercio','9.2. Inhabilitado para ejercer comercio?'),
            ('p3_mora_obligaciones','9.3. En mora con Estado/SuperBancos?'),
            ('p4_vinculo_financiero','9.4. Vinculo con sistema financiero?'),
            ('p5_funcionario_iess','9.5. Funcionario IESS/BIESS?'),
            ('p6_interes_aseguradoras','9.6. Interes en aseguradoras?'),
            ('p7_removido_organismo','9.7. Removido por Organismo de Control (5 años)?'),
            ('p8_sri','9.8. Obligaciones pendientes SRI?'),
            ('p9_castigo_financiero','9.9. Castigo financiero (5 años)?'),
            ('p10_litigio_iess','9.10. Litigando contra IESS/BIESS?'),
            ('p11_procesado_corrupcion','9.11. Procesado por corrupcion?'),
            ('p12_contraloria','9.12. Responsabilidades Contraloria?'),
            ('p13_uafe','9.13. Registrado en UAFE?'),
        ]:
            inh_rows += f'<tr><td>{etiqueta}</td><td class="r">{si_no(getattr(inhabilidades, campo, None))}</td></tr>'

    fam_rows  = ''.join(f'<tr><td>{f.nombres}</td><td>{f.get_parentesco_display()}</td><td>{f.get_institucion_display()}</td><td>{f.area}</td><td>{f.cargo}</td></tr>' for f in familiares)
    form_rows = ''.join(f'<tr><td>{f.get_nivel_display()}</td><td>{f.get_tipo_display()}</td><td>{f.titulo}</td><td>{f.institucion}</td><td>{f.get_area_estudios_display()}</td><td>{f.num_senescyt}</td><td>{f.fecha_senescyt.strftime("%d/%m/%Y")}</td></tr>' for f in formaciones)
    exp_rows  = ''.join(f'<tr><td>{e.get_tipo_display()}</td><td>{e.cargo}</td><td>{e.institucion}</td><td>{e.get_actividades_area_display()}</td><td>{e.fecha_inicio.strftime("%d/%m/%Y")}</td><td>{e.fecha_fin.strftime("%d/%m/%Y")}</td><td>{e.tiempo_calculado["texto"]}</td></tr>' for e in experiencias)
    cap_rows  = ''.join(f'<tr><td>{c.get_tipo_evento_display()}</td><td>{c.nombre}</td><td>{c.institucion}</td><td>{c.fecha_inicio.strftime("%d/%m/%Y")}</td><td>{c.fecha_fin.strftime("%d/%m/%Y")}</td><td>{c.horas}h</td></tr>' for c in capacitaciones)
    pub_rows  = ''.join(f'<tr><td>{p.titulo}</td><td>{p.get_tipo_display()}</td><td>{p.medio}</td><td>{p.fecha.strftime("%d/%m/%Y")}</td></tr>' for p in publicaciones)

    html = f'''<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<style>*{{margin:0;padding:0;box-sizing:border-box;}}body{{font-family:Helvetica,Arial,sans-serif;font-size:10px;color:#222;}}
.header{{background-color:#003580;color:white;padding:12px 20px;margin-bottom:16px;}}
.header h1{{font-size:14px;margin-bottom:2px;}}.header p{{font-size:9px;}}
.codigo{{font-size:12px;font-weight:bold;color:#003580;margin-bottom:14px;padding:6px 10px;border:2px solid #003580;display:inline-block;}}
h2{{font-size:11px;color:#003580;border-bottom:1px solid #003580;padding-bottom:3px;margin:14px 0 8px 0;text-transform:uppercase;}}
.grid{{width:100%;margin-bottom:8px;}}.grid td{{padding:3px 6px;font-size:9px;vertical-align:top;width:33%;}}
.lbl{{color:#666;font-size:8px;display:block;margin-bottom:1px;}}.val{{font-weight:bold;}}
table.t{{width:100%;border-collapse:collapse;margin-bottom:10px;font-size:9px;}}
table.t th{{background-color:#003580;color:white;padding:4px 6px;text-align:left;font-size:8px;}}
table.t td{{padding:3px 6px;border-bottom:1px solid #ddd;}}
.inh{{width:100%;border-collapse:collapse;font-size:9px;}}
.inh td{{padding:3px 6px;border-bottom:1px solid #eee;}}.inh .r{{font-weight:bold;width:25px;text-align:center;}}
.decl{{background-color:#f8f8f8;border:1px solid #ddd;padding:10px 12px;margin:12px 0;font-size:8px;line-height:1.6;}}
.decl p{{margin-bottom:4px;}}.nd{{color:#999;font-style:italic;font-size:9px;margin-bottom:8px;}}
</style></head><body>
<div class="header"><h1>INSTITUTO ECUATORIANO DE SEGURIDAD SOCIAL</h1>
<p>Concurso Publico de Meritos y Oposicion - Directorio BIESS</p></div>
<div class="codigo">Codigo: {postulante.codigo_unico} | Fecha: {postulante.creado_en.strftime("%d/%m/%Y")}</div>
<h2>1. Sector</h2><table class="grid"><tr>
<td colspan="2"><span class="lbl">Cargo</span><span class="val">{postulante.get_sector_display()}</span></td>
<td><span class="lbl">Organizacion</span><span class="val">{postulante.nombre_organizacion or "No"}</span></td>
</tr></table>
<h2>2. Informacion Personal</h2><table class="grid"><tr>
<td><span class="lbl">Nombres</span><span class="val">{postulante.nombres}</span></td>
<td><span class="lbl">Apellidos</span><span class="val">{postulante.apellidos}</span></td>
<td><span class="lbl">Cedula</span><span class="val">{postulante.usuario.cedula}</span></td>
</tr><tr>
<td><span class="lbl">Genero</span><span class="val">{postulante.get_genero_display()}</span></td>
<td><span class="lbl">Nacimiento</span><span class="val">{postulante.fecha_nacimiento.strftime("%d/%m/%Y") if postulante.fecha_nacimiento else "-"}</span></td>
<td><span class="lbl">Nacionalidad</span><span class="val">{postulante.nacionalidad}</span></td>
</tr><tr>
<td><span class="lbl">Estado civil</span><span class="val">{postulante.get_estado_civil_display()}</span></td>
<td><span class="lbl">Conyuge</span><span class="val">{postulante.conyuge_nombres or "No aplica"}</span></td>
<td><span class="lbl">Cedula conyuge</span><span class="val">{postulante.conyuge_cedula or "No aplica"}</span></td>
</tr></table>
<h2>3. Domicilio</h2><table class="grid"><tr>
<td><span class="lbl">Pais</span><span class="val">{postulante.pais}</span></td>
<td><span class="lbl">Provincia</span><span class="val">{postulante.provincia}</span></td>
<td><span class="lbl">Ciudad</span><span class="val">{postulante.ciudad}</span></td>
</tr><tr>
<td><span class="lbl">Calle principal</span><span class="val">{postulante.calle_principal} {postulante.numero}</span></td>
<td><span class="lbl">Calle secundaria</span><span class="val">{postulante.calle_secundaria}</span></td>
<td><span class="lbl">Sector</span><span class="val">{postulante.sector_domicilio}</span></td>
</tr><tr>
<td><span class="lbl">Celular</span><span class="val">{postulante.telefono_celular}</span></td>
<td><span class="lbl">Tel. domicilio</span><span class="val">{postulante.telefono_domicilio or "No registra"}</span></td>
<td><span class="lbl">Correo</span><span class="val">{postulante.usuario.email}</span></td>
</tr></table>
<h2>4. Familiares IESS/BIESS</h2>
{"<table class='t'><thead><tr><th>Nombres</th><th>Parentesco</th><th>Institucion</th><th>Area</th><th>Cargo</th></tr></thead><tbody>" + fam_rows + "</tbody></table>" if fam_rows else "<p class='nd'>No registra.</p>"}
<h2>5. Formacion Academica</h2>
{"<table class='t'><thead><tr><th>Nivel</th><th>Tipo</th><th>Titulo</th><th>Institucion</th><th>Area</th><th>SENESCYT</th><th>Fecha</th></tr></thead><tbody>" + form_rows + "</tbody></table>" if form_rows else "<p class='nd'>No registra.</p>"}
<h2>6. Experiencia Profesional</h2>
{"<table class='t'><thead><tr><th>Tipo</th><th>Cargo</th><th>Institucion</th><th>Area</th><th>Inicio</th><th>Fin</th><th>Tiempo</th></tr></thead><tbody>" + exp_rows + "</tbody></table>" if exp_rows else "<p class='nd'>No registra.</p>"}
<h2>7. Capacitacion</h2>
{"<table class='t'><thead><tr><th>Tipo</th><th>Nombre</th><th>Institucion</th><th>Inicio</th><th>Fin</th><th>Horas</th></tr></thead><tbody>" + cap_rows + "</tbody></table>" if cap_rows else "<p class='nd'>No registra.</p>"}
<h2>8. Publicaciones</h2>
{"<table class='t'><thead><tr><th>Titulo</th><th>Tipo</th><th>Medio</th><th>Fecha</th></tr></thead><tbody>" + pub_rows + "</tbody></table>" if pub_rows else "<p class='nd'>No registra.</p>"}
<h2>9. Inhabilidades</h2>
{"<table class='inh'>" + inh_rows + "</table>" if inh_rows else "<p class='nd'>No registra.</p>"}
<div class="decl"><p><strong>Declaracion:</strong></p>
<p>1. Autorizo al IESS el uso de mis datos personales para este concurso.</p>
<p>2. Declaro que la informacion es veridica.</p>
<p>3. Acepto las normas del concurso y mi posible descalificacion por incumplimiento.</p>
<p>4. Acepto la publicacion de mi nombre en la pagina web del IESS/BIESS.</p></div>
<table width="100%" style="margin-top:20px;"><tr><td width="40%"></td>
<td width="20%" align="center"><br/><br/><br/><br/><br/>
<div style="border-top:1px solid #333;padding-top:6px;font-size:9px;">
{postulante.nombres} {postulante.apellidos}<br/>C.I.: {postulante.cedula}<br/>Firma del Postulante
</div></td><td width="40%"></td></tr></table>
</body></html>'''

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="postulacion_{postulante.codigo_unico}.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response


def solo_staff(view_func):
    from functools import wraps
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or not request.user.is_staff:
            return redirect('login')
        return view_func(request, *args, **kwargs)
    return wrapper


@solo_staff
def admin_postulaciones(request):
    sector = request.GET.get('sector', '')
    estado = request.GET.get('estado', '')
    buscar = request.GET.get('buscar', '')
    postulantes = Postulante.objects.select_related('usuario').all().order_by('-creado_en')
    if sector:
        postulantes = postulantes.filter(sector=sector)
    if estado:
        postulantes = postulantes.filter(estado=estado)
    if buscar:
        postulantes = postulantes.filter(cedula__icontains=buscar) | \
                      postulantes.filter(nombres__icontains=buscar) | \
                      postulantes.filter(apellidos__icontains=buscar)
    return render(request, 'postulaciones/admin_postulaciones.html', {
        'postulantes':       postulantes,
        'total':             Postulante.objects.count(),
        'total_afiliados':   Postulante.objects.filter(sector='AFILIADO').count(),
        'total_jubilados':   Postulante.objects.filter(sector='JUBILADO').count(),
        'total_empleadores': Postulante.objects.filter(sector='EMPLEADOR').count(),
        'total_enviados':    Postulante.objects.filter(estado='ENVIADO').count(),
        'total_borradores':  Postulante.objects.filter(estado='BORRADOR').count(),
        'sector_actual':     sector,
        'estado_actual':     estado,
        'buscar':            buscar,
        'SECTOR_CHOICES':    Postulante.SECTOR_CHOICES,
        'ESTADO_CHOICES':    Postulante.ESTADO_CHOICES,
    })


@solo_staff
def admin_detalle_postulante(request, pk):
    postulante = get_object_or_404(Postulante, pk=pk)
    return render(request, 'postulaciones/admin_detalle.html', {
        'postulante':     postulante,
        'familiares':     postulante.familiares.all(),
        'formaciones':    postulante.formaciones.all(),
        'experiencias':   postulante.experiencias.all(),
        'capacitaciones': postulante.capacitaciones.all(),
        'publicaciones':  postulante.publicaciones.all(),
        'inhabilidades':  getattr(postulante, 'inhabilidades', None),
    })


@solo_staff
def admin_cambiar_estado(request, pk):
    if request.method == 'POST':
        postulante   = get_object_or_404(Postulante, pk=pk)
        nuevo_estado = request.POST.get('estado')
        if nuevo_estado:
            postulante.estado = nuevo_estado
            postulante.save()
            notificar_cambio_estado(postulante)
            messages.success(request, f'Estado actualizado a {postulante.get_estado_display()}.')
    return redirect('admin_postulaciones')


@solo_staff
def exportar_excel(request):
    postulantes = Postulante.objects.select_related('usuario').all().order_by('-creado_en')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Postulaciones'
    header_font  = Font(bold=True, color='FFFFFF', size=10)
    header_fill  = PatternFill(start_color='003580', end_color='003580', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center')
    headers = ['N°','Código','Cédula','Apellidos','Nombres','Sector','Estado','Correo',
               'Celular','País','Provincia','Ciudad','Formaciones','Experiencias','Capacitaciones','Fecha registro']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align
    ws.row_dimensions[1].height = 30
    for idx, p in enumerate(postulantes, 1):
        fila = [idx, p.codigo_unico, p.cedula, p.apellidos, p.nombres,
                p.get_sector_display(), p.get_estado_display(), p.usuario.email,
                p.telefono_celular, p.pais, p.provincia, p.ciudad,
                p.formaciones.count(), p.experiencias.count(), p.capacitaciones.count(),
                p.creado_en.strftime('%d/%m/%Y %H:%M')]
        for col, valor in enumerate(fila, 1):
            cell = ws.cell(row=idx + 1, column=col, value=valor); cell.alignment = center_align
    anchos = [5,15,12,20,20,25,15,30,13,12,15,15,12,12,12,18]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho
    ws2 = wb.create_sheet('Resumen')
    ws2.append(['Sector', 'Total'])
    for row in Postulante.objects.values('sector').annotate(total=Count('id')):
        ws2.append([dict(Postulante.SECTOR_CHOICES).get(row['sector'], row['sector']), row['total']])
    ws2.append([])
    ws2.append(['Estado', 'Total'])
    for row in Postulante.objects.values('estado').annotate(total=Count('id')):
        ws2.append([dict(Postulante.ESTADO_CHOICES).get(row['estado'], row['estado']), row['total']])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="postulaciones_biess.xlsx"'
    wb.save(response)
    return response


@solo_staff
def admin_pdf_postulante(request, pk):
    postulante     = get_object_or_404(Postulante, pk=pk)
    familiares     = postulante.familiares.all()
    formaciones    = postulante.formaciones.all()
    experiencias   = postulante.experiencias.all()
    capacitaciones = postulante.capacitaciones.all()
    publicaciones  = postulante.publicaciones.all()
    inhabilidades  = getattr(postulante, 'inhabilidades', None)

    def si_no(val):
        if val is True: return 'Si'
        elif val is False: return 'No'
        return '-'

    inh_rows = ''
    if inhabilidades:
        for campo, etiqueta in [
            ('p1_goce_derechos','1. Ecuatoriano/a en pleno goce de derechos?'),
            ('p2_inhabilitado_comercio','2. Inhabilitado para ejercer el comercio?'),
            ('p3_mora_obligaciones','3. En mora con el Estado o SuperBancos?'),
            ('p4_vinculo_financiero','4. Vinculo con sistema financiero?'),
            ('p5_funcionario_iess','5. Funcionario IESS/BIESS?'),
            ('p6_interes_aseguradoras','6. Interes en aseguradoras?'),
            ('p7_removido_organismo','7. Removido por Organismo de Control (5 años)?'),
            ('p8_sri','8. Obligaciones pendientes con SRI?'),
            ('p9_castigo_financiero','9. Castigo financiero (5 años)?'),
            ('p10_litigio_iess','10. Litigando contra IESS/BIESS?'),
            ('p11_procesado_corrupcion','11. Procesado por corrupcion?'),
            ('p12_contraloria','12. Responsabilidades por Contraloria?'),
            ('p13_uafe','13. Registrado en UAFE?'),
        ]:
            inh_rows += f'<tr><td>{etiqueta}</td><td class="r">{si_no(getattr(inhabilidades, campo, None))}</td></tr>'

    fam_rows  = ''.join(f'<tr><td>{f.nombres}</td><td>{f.get_parentesco_display()}</td><td>{f.get_institucion_display()}</td><td>{f.area}</td><td>{f.cargo}</td></tr>' for f in familiares)
    form_rows = ''.join(f'<tr><td>{f.get_nivel_display()}</td><td>{f.titulo}</td><td>{f.institucion}</td><td>{f.num_senescyt}</td><td>{f.fecha_senescyt.strftime("%d/%m/%Y")}</td></tr>' for f in formaciones)
    exp_rows  = ''.join(f'<tr><td>{e.cargo}</td><td>{e.institucion}</td><td>{e.fecha_inicio.strftime("%d/%m/%Y")}</td><td>{e.fecha_fin.strftime("%d/%m/%Y")}</td><td>{e.tiempo_calculado["texto"]}</td></tr>' for e in experiencias)
    cap_rows  = ''.join(f'<tr><td>{c.nombre}</td><td>{c.institucion}</td><td>{c.horas}h</td></tr>' for c in capacitaciones)
    pub_rows  = ''.join(f'<tr><td>{p.titulo}</td><td>{p.get_tipo_display()}</td><td>{p.fecha.strftime("%d/%m/%Y")}</td></tr>' for p in publicaciones)

    html = f'''<!DOCTYPE html><html><head><meta charset="UTF-8">
<style>*{{margin:0;padding:0;}}body{{font-family:Helvetica;font-size:10px;}}
.h{{background:#003580;color:white;padding:10px;margin-bottom:12px;}}
h2{{font-size:10px;color:#003580;border-bottom:1px solid #003580;margin:10px 0 6px;}}
table{{width:100%;border-collapse:collapse;font-size:9px;margin-bottom:8px;}}
th{{background:#003580;color:white;padding:3px 5px;text-align:left;}}
td{{padding:3px 5px;border-bottom:1px solid #eee;}}.r{{font-weight:bold;width:20px;}}
</style></head><body>
<div class="h"><b>IESS — Concurso Publico Directorio BIESS</b><br/>
Codigo: {postulante.codigo_unico} | {postulante.nombres} {postulante.apellidos} | C.I.: {postulante.usuario.cedula}</div>
<h2>Sector: {postulante.get_sector_display()}</h2>
<h2>Familiares IESS/BIESS</h2>
{"<table><tr><th>Nombres</th><th>Parentesco</th><th>Institucion</th><th>Area</th><th>Cargo</th></tr>" + fam_rows + "</table>" if fam_rows else "<p>No registra.</p>"}
<h2>Formacion Academica</h2>
{"<table><tr><th>Nivel</th><th>Titulo</th><th>Institucion</th><th>SENESCYT</th><th>Fecha</th></tr>" + form_rows + "</table>" if form_rows else "<p>No registra.</p>"}
<h2>Experiencia Profesional</h2>
{"<table><tr><th>Cargo</th><th>Institucion</th><th>Inicio</th><th>Fin</th><th>Tiempo</th></tr>" + exp_rows + "</table>" if exp_rows else "<p>No registra.</p>"}
<h2>Capacitacion</h2>
{"<table><tr><th>Nombre</th><th>Institucion</th><th>Horas</th></tr>" + cap_rows + "</table>" if cap_rows else "<p>No registra.</p>"}
<h2>Publicaciones</h2>
{"<table><tr><th>Titulo</th><th>Tipo</th><th>Fecha</th></tr>" + pub_rows + "</table>" if pub_rows else "<p>No registra.</p>"}
<h2>Inhabilidades</h2>
{"<table>" + inh_rows + "</table>" if inh_rows else "<p>No registra.</p>"}
</body></html>'''

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="postulacion_{postulante.codigo_unico}.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response
